import openpyxl
import re
import json
import os
import uuid
from datetime import datetime

# Source IDs
Y2_SOURCE_ID = "b81a2f45-3129-5f21-a189-9a2c14022a8e"
Y3_SOURCE_ID = "7c92e104-58ab-590d-b21a-4d1f851003ef"

# Standard time mapping for normalized 24-hr format
TIME_MAPPINGS = {
    # Morning
    "8:15 to 9:05": ("08:15", "09:05"),
    "8.15-9.05": ("08:15", "09:05"),
    "8:15-9:05": ("08:15", "09:05"),
    
    "9:05 to 9:55": ("09:05", "09:55"),
    "9.05-9.55": ("09:05", "09:55"),
    "9:05-9:55": ("09:05", "09:55"),
    
    "9:55 to 10:45": ("09:55", "10:45"),
    "9.55-10.45": ("09:55", "10:45"),
    "9:55-10:45": ("09:55", "10:45"),
    
    # Morning Break
    "10:45 to 11:00": ("10:45", "11:00"),
    "10.45-11.00": ("10:45", "11:00"),
    "10:45-11:00": ("10:45", "11:00"),
    
    # Mid-day
    "11:00 to 11:50": ("11:00", "11:50"),
    "11.00.-11.50": ("11:00", "11:50"),
    "11.00-11.50": ("11:00", "11:50"),
    
    "11:50 to 12:40": ("11:50", "12:40"),
    "11.50-12.40": ("11:50", "12:40"),
    
    "12:40 to 1:30": ("12:40", "13:30"),
    "12.40-1.30": ("12:40", "13:30"),
    "12:40-1:30": ("12:40", "13:30"),
    
    "1:30 to 2:30": ("13:30", "14:30"),
    "1.30-2.20": ("13:30", "14:20"),
    "1:30-2:20": ("13:30", "14:20"),
    
    "2.20-2.30": ("14:20", "14:30"),
    "2:20-2:30": ("14:20", "14:30"),
    
    "2:30 to 3:20": ("14:30", "15:20"),
    "2.20-3.10": ("14:20", "15:10"),
    "2:20-3:10": ("14:20", "15:10"),
    "2:30-3:20": ("14:30", "15:20"),
    
    "3:20 to 4:10": ("15:20", "16:10"),
    "3.10-4.00": ("15:10", "16:00"),
    "3:10-4:00": ("15:10", "16:00"),
}

DAY_MAP = {
    'MON': 'Monday', 'MONDAY': 'Monday',
    'TUE': 'Tuesday', 'TUESDAY': 'Tuesday',
    'WED': 'Wednesday', 'WEDNESDAY': 'Wednesday',
    'THU': 'Thursday', 'THURSDAY': 'Thursday',
    'FRI': 'Friday', 'FRIDAY': 'Friday',
    'SAT': 'Saturday', 'SATURDAY': 'Saturday'
}

def clean(val):
    if val is None:
        return ""
    return str(val).strip()

def parse_time_slot(raw_str):
    raw = raw_str.strip()
    if raw in TIME_MAPPINGS:
        return TIME_MAPPINGS[raw]
    
    # Fallback pattern matching
    m = re.search(r'(\d{1,2})[\.:](\d{2})\s*(?:to|-)\s*(\d{1,2})[\.:](\d{2})', raw)
    if m:
        sh, sm, eh, em = map(int, m.groups())
        if sh < 8:
            sh += 12
        if eh < 8 or (eh < sh and eh != 12):
            eh += 12
        return f"{sh:02d}:{sm:02d}", f"{eh:02d}:{em:02d}"
    return None, None

def parse_class_type(text):
    text_u = text.upper()
    if '[L]' in text_u or '(L)' in text_u:
        return 'Lecture'
    elif '[T]' in text_u or '(T)' in text_u or '[T}' in text_u:
        return 'Tutorial'
    elif '[P]' in text_u or '(P)' in text_u:
        return 'Practical'
    elif 'BREAK' in text_u:
        return 'Break'
    elif 'SELF LEARNING' in text_u or 'ADVANCED LEARNING' in text_u:
        return 'Self Learning'
    elif 'COUN' in text_u:
        return 'Counselling'
    elif 'EXPERIENTIAL' in text_u or 'PROJECT' in text_u:
        return 'Experiential Learning'
    elif 'TRAINING' in text_u:
        return 'Training'
    elif 'SPORTS' in text_u or 'LIB' in text_u or 'LIBRARY' in text_u or 'NPTEL' in text_u or 'MINORS' in text_u or 'CC' in text_u:
        return 'Co-Curricular / Activity'
    return 'Lecture'

def extract_room_and_code(cell_text, default_room):
    text = cell_text.strip()
    if not text:
        return None, None, None
        
    if text.upper() == 'BREAK':
        return 'BREAK', None, None
        
    # Check for room explicitly enclosed in brackets/parentheses e.g. [N-314A], (C2-LAB-3), [N-411], (N-407)
    m_room = re.search(r'[\[\(]([N|C|A|B][-\w\d]+)[\]\)]', text)
    explicit_room = None
    if m_room:
        explicit_room = m_room.group(1).strip()

    # Clean subject code
    code_text = text
    if explicit_room:
        code_text = code_text.replace(f"[{explicit_room}]", "").replace(f"({explicit_room})", "").replace(explicit_room, "").strip()
        
    # Remove class type bracket e.g. [L], [T], (P), [T}
    code_text = re.sub(r'[\[\(\{][L|T|P|S|T\}][\]\)\}]', '', code_text).strip()
    code_text = re.sub(r'\s+', ' ', code_text).strip()
    
    final_room = explicit_room if explicit_room else default_room
    return code_text, final_room, explicit_room


def parse_section_legend(ws):
    fac_map = {}
    for r in range(9, ws.max_row + 1):
        c2 = clean(ws.cell(row=r, column=2).value)
        c_fac = clean(ws.cell(row=r, column=6).value) or clean(ws.cell(row=r, column=5).value) or clean(ws.cell(row=r, column=4).value)
        if c2 and c_fac:
            fname = re.sub(r'\(\d+\)', '', c_fac).strip()
            fname = re.sub(r'\s+', ' ', fname).strip()
            if not fname:
                continue
                
            code_m = re.search(r'\(([A-Z0-9\s\-]+)\)$', c2) or re.search(r'\b([A-Z]{2,6})\b', c2)
            if code_m:
                code_key = code_m.group(1).strip().upper()
                fac_map[code_key] = fname
            
            c2_u = c2.upper()
            if 'COMPUTING ETHICS' in c2_u or 'CE' in c2_u:
                fac_map['CE'] = fname
                fac_map['CE -'] = fname
            elif 'MACHINE LEARNING' in c2_u or 'ML' in c2_u:
                fac_map['ML'] = fname
                fac_map['ML -'] = fname
            elif 'COMPUTER NETWORKS' in c2_u or 'CN' in c2_u:
                fac_map['CN'] = fname
                fac_map['CN -'] = fname
            elif 'OPTIMIZATION TECHNIQUES' in c2_u or 'OT' in c2_u:
                fac_map['OT'] = fname
            elif 'MODERN FRONT-END' in c2_u or 'MFEF' in c2_u:
                fac_map['MFEF'] = fname
            elif 'DATA STRUCTURES' in c2_u or 'DS' in c2_u:
                fac_map['DS'] = fname
            elif 'DATABASE' in c2_u or 'DBMS' in c2_u:
                fac_map['DBMS'] = fname
            elif 'OBJECT ORIENTED' in c2_u or 'OOPS' in c2_u:
                fac_map['OOPS'] = fname
            elif 'DIGITAL LOGIC' in c2_u or 'DLD' in c2_u:
                fac_map['DLD'] = fname
            elif 'ARTIFICIAL INTELLIGENCE' in c2_u or 'AI' in c2_u:
                fac_map['AI'] = fname
            elif 'DISCRETE' in c2_u or 'DMS' in c2_u:
                fac_map['DMS'] = fname
    return fac_map


def process_workbook(file_path, year, source_id):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    records = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        merged_ranges = ws.merged_cells.ranges
        
        def get_merged_extent(r, c):
            for rng in merged_ranges:
                if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
                    return rng.min_row, rng.max_row, rng.min_col, rng.max_col
            return r, r, c, c

        fac_map = parse_section_legend(ws)
        
        top_text = ""
        for r in range(1, 6):
            row_str = " ".join([clean(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)])
            top_text += " " + row_str
            
        m_sec = re.search(r'SECTION[-:\s]*(\d+)', top_text, re.IGNORECASE)
        sec_num = m_sec.group(1).strip() if m_sec else sheet_name.replace('sec', '')
        
        m_room = re.search(r'\[(N-\d+[A-Z]?)\]', top_text)
        if not m_room:
            m_room = re.search(r'ROOM\s*(?:NO)?[-:\s]*([N|C][-\w\d]+)', top_text, re.IGNORECASE)
        default_room = m_room.group(1).strip() if m_room else ("N-313" if year == 2 else "N-407")
        
        time_r = None
        day_c = None
        for r in range(1, min(15, ws.max_row + 1)):
            for c in range(1, min(6, ws.max_column + 1)):
                val = clean(ws.cell(row=r, column=c).value)
                if val.lower() == 'day':
                    time_r = r
                    day_c = c
                    break
            if time_r:
                break
                
        if not time_r:
            continue
            
        time_cols = []
        col_map = {}
        for c in range(day_c + 1, ws.max_column + 1):
            t_val = clean(ws.cell(row=time_r, column=c).value)
            if t_val:
                st, et = parse_time_slot(t_val)
                if st and et:
                    time_cols.append((c, t_val, st, et))
                    col_map[c] = (st, et)
                    
        for r in range(time_r + 1, time_r + 10):
            d_val = clean(ws.cell(row=r, column=day_c).value).upper()
            if d_val in DAY_MAP:
                day_name = DAY_MAP[d_val]
                processed_cols = set()
                
                for c_idx, raw_time, start_time, end_time in time_cols:
                    if c_idx in processed_cols:
                        continue
                        
                    r_min, r_max, c_min, c_max = get_merged_extent(r, c_idx)
                    cell_val = clean(ws.cell(row=r, column=c_min).value)
                    
                    if not cell_val:
                        processed_cols.add(c_idx)
                        continue

                    c_start = c_min
                    c_end = min(c_max, max(col_map.keys()))
                    while c_end not in col_map and c_end > c_start:
                        c_end -= 1

                    real_st = col_map[c_start][0] if c_start in col_map else start_time
                    real_et = col_map[c_end][1] if c_end in col_map else end_time

                    for mark_c in range(c_start, c_max + 1):
                        processed_cols.add(mark_c)

                    code, room, explicit_room = extract_room_and_code(cell_val, default_room)
                    class_type = parse_class_type(cell_val)
                    
                    code_clean = (code or cell_val).strip().upper()
                    faculty_name = fac_map.get(code_clean) or fac_map.get(code_clean.replace(" -", "")) or fac_map.get(code_clean.split()[0])
                    
                    rec = {
                        "academic_year": "2024-2025",
                        "project_target_academic_year": "2026-2027",
                        "project_usage": "current_student_timetable",
                        "year": year,
                        "section": sec_num,
                        "day": day_name,
                        "start_time": real_st,
                        "end_time": real_et,
                        "subject_code": code or cell_val,
                        "subject_name": code or cell_val,
                        "class_type": class_type,
                        "room": room,
                        "section_default_room": default_room,
                        "faculty": faculty_name,
                        "source_id": source_id,
                        "confidence": "high",
                        "last_verified": "2026-08-27T00:00:00Z"
                    }
                    records.append(rec)
                        
    return records


def process_y2_workbook(file_path):
    return process_workbook(file_path, 2, Y2_SOURCE_ID)


def process_y3_workbook(file_path):
    return process_workbook(file_path, 3, Y3_SOURCE_ID)

def update_sources_json():
    sources_path = "database/extracted/sources.json"
    with open(sources_path, "r", encoding="utf-8") as f:
        sources = json.load(f)
        
    existing_ids = {s["id"] for s in sources}
    
    y2_src = {
        "id": Y2_SOURCE_ID,
        "source_type": "official_document",
        "source_name": "VIGNAN CSE II Year II Semester Official Timetable Workbook",
        "document_name": "2nd yt TT (2) (1) (2).xlsx",
        "source_url": "database/raw/timetables/2nd yt TT (2) (1) (2).xlsx",
        "verified_at": "2026-08-27T00:00:00Z",
        "description": "Official section-wise timetable workbook for II B.Tech CSE (Sections 1 to 19)"
    }
    
    y3_src = {
        "id": Y3_SOURCE_ID,
        "source_type": "official_document",
        "source_name": "VIGNAN CSE III Year II Semester Official Timetable Spreadsheet",
        "document_name": "year3_cse_timetable_2026.xlsx",
        "source_url": "https://tinyurl.com/3rdyearTT2026",
        "verified_at": "2026-08-27T00:00:00Z",
        "description": "Official section-wise timetable spreadsheet for III B.Tech CSE (Sections 1 to 22)"
    }
    
    if Y2_SOURCE_ID not in existing_ids:
        sources.append(y2_src)
    if Y3_SOURCE_ID not in existing_ids:
        sources.append(y3_src)
        
    with open(sources_path, "w", encoding="utf-8") as f:
        json.dump(sources, f, indent=2)
    print("Updated database/extracted/sources.json")

def main():
    os.makedirs("database/extracted/timetables", exist_ok=True)
    
    update_sources_json()
    
    print("Processing Year 2 workbook...")
    y2_records = process_y2_workbook("database/raw/timetables/2nd yt TT (2) (1) (2).xlsx")
    with open("database/extracted/timetables/year2_timetable.json", "w", encoding="utf-8") as f:
        json.dump(y2_records, f, indent=2)
    print(f"Saved {len(y2_records)} Year 2 records to database/extracted/timetables/year2_timetable.json")
    
    print("Processing Year 3 workbook...")
    y3_records = process_y3_workbook("database/raw/timetables/year3_cse_timetable_2026.xlsx")
    with open("database/extracted/timetables/year3_timetable.json", "w", encoding="utf-8") as f:
        json.dump(y3_records, f, indent=2)
    print(f"Saved {len(y3_records)} Year 3 records to database/extracted/timetables/year3_timetable.json")

if __name__ == "__main__":
    main()
